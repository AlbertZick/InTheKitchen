import os
import sys
import re
from  openpyxl import load_workbook
import math


class Define():
   # kinds of Port connection
   _OUTPUT = 0
   _INPUT  = 1
   _PARAM  = 2
   # kinds of Connect type
   _FLOAT  = 0
   _SPEC_NAME = 1
   _PRE_DEC = 2
   _SAME_NAME = 3

   # Column definition
   _DIRECTION  = 1
   _TYPE       = 2
   _ARRAY      = 3
   _NAME       = 4
   _INSTNAME   = 5

   _UNKNOWN = 99

   ModuleDef = '''//===========================================================================
// Author : 
// Module : [name]
//===========================================================================

module [name] #([param]) (
[portLst]
) ;
'''

   WireDef = '''
//=======START DECLARING WIRES ================================================//
[WireLst]
//=======FINISH DECLARING WIRES ===============================================//
'''
   ModuleInst = '''
[name] #([param]) [instName] (
[portLst]
) ;
'''

   EndDef = '''

endmodule
'''


class Port():
   """docstring for Port"""
   def __init__(self):
      self.direction = Define._UNKNOWN
      self.type      = ''
      self.array     = ''
      self.name      = ''
      self.connectType = Define._UNKNOWN
      self.InstName  = ''
      self.row       = -1


class Module():
   def __init__(self):
      self.name               = ''
      self.InstName           = ''
      self.LstPort            = []

   def findInstNameByRow (self, row):
      for e in self.LstPort:
         if e.row == row \
            and (e.connectType == Define._SPEC_NAME \
                  or e.connectType == Define._SAME_NAME):
            return e.InstName
      
      return '~~'

   def findPortNameByRow (self, row):
      for e in self.LstPort:
         if e.row == row \
            and (e.connectType == Define._SPEC_NAME \
                 or e.connectType == Define._SAME_NAME):
            return e.name
      
      return '~~'

class LstModule():
   def __init__(self):
      self.LstModule          = []
      self.LstWireDeClaration = []
      self.LstTopModulePort = []

   def findInstNameByRow (self, row, module_i):
      for i in range(module_i, -1, -1):
         e = self.LstModule[i]
         InstName = e.findInstNameByRow(row)
         if InstName != '~~':
            return InstName

      return '~~'

   def findPortNameByRow (self, row, module_i):
      for i in range(module_i, -1, -1):
         e = self.LstModule[i]
         InstName = e.findPortNameByRow(row)
         if InstName != '~~':
            return InstName

      return '~~'

   def containInstNameInWireLst(self, InstName):
      for wirePort in self.LstWireDeClaration:
         if InstName == wirePort.InstName:
            return True

      return False

   def containInstNameInTopModulePortLst (self, InstName):
      for port in self.LstTopModulePort:
         if InstName == port.name:
            return True

      return False


class XLParser(object):
   def __init__(self, FileName, SheetName=None):
      self.WB = load_workbook(filename=FileName)
      self.LstWS = {}
      if bool(SheetName):
         self.LstWS['SheetName'] = self.WB[SheetName]
      else:
         pass

   def saveXLFile(self):
      self.WB.save(filename=FileName)

   def getLstSheet (self):
      return self.WB.sheetnames

   def findNumberOfModuleInSheet (self, SheetName):
      WS = self.WB[SheetName]
      row = 1
      Cntr = 0
      col = 1
      while (col <= WS.max_column):
         value = WS.cell(row=row, column=col).value
         
         if not value:
            col += 1
         elif value.strip() == '':
            col += 1
         else:
            col = (math.floor((col-1)/5)+1)*5 + 1
            Cntr += 1

      return Cntr

   def parseXL(self, SheetName):
      WS = self.WB[SheetName]
      NumModule = self.findNumberOfModuleInSheet(SheetName)
      m_LstModule = LstModule()
      baseCol = -5


      for i in range(NumModule):
         baseCol += 5
         m_module = Module()
         # get module name
         m_module.name = str(WS.cell(row=1, column=baseCol+Define._NAME).value).strip().replace(' ', '')
         m_module.InstName = str(WS.cell(row=1, column=baseCol+Define._INSTNAME).value).strip().replace(' ', '')
         for row in range(1, WS.max_row+1):
            # check port direction
            direction = WS.cell(row=row, column=baseCol+Define._DIRECTION).value
            if not direction:
               continue
            direction = str(direction).strip().replace(' ', '').lower()
            port = Port()
            if   direction == 'i':
               port.direction = Define._INPUT
            elif direction == 'o':
               port.direction = Define._OUTPUT
            elif direction == 'p':
               port.direction = Define._PARAM
            else:
               continue

            # check port type
            PortType = WS.cell(row=row, column=baseCol+Define._TYPE).value
            PortType = '' if not PortType else PortType
            if '\n' in PortType or ' ' in PortType:
               PortType = PortType.strip().replace(' ', '')
               WS.cell(row=row, column=baseCol+Define._TYPE, value=PortType)

            port.type = PortType
            # check name
            Name = WS.cell(row=row, column=baseCol+Define._NAME).value
            Name = '' if not Name else Name
            if '\n' in Name or ' ' in Name:
               Name = Name.strip().replace(' ', '')
               WS.cell(row=row, column=baseCol+Define._NAME, value=Name)
            if bool(re.search(r'\[\d{1,}:\d{1,}\]', Name)):
               openBrIdx = Name.index('[')
               closeBrIdx = Name.index(']')+1
               tmp = list(Name)
               Array = ''.join(tmp[openBrIdx:closeBrIdx])
               tmp[openBrIdx:closeBrIdx] = ''
               Name = ''.join(tmp)
               WS.cell(row=row, column=baseCol+Define._NAME, value=Name)
               WS.cell(row=row, column=baseCol+Define._ARRAY, value=Array)

            port.name = Name

            # check array
            Array = WS.cell(row=row, column=baseCol+Define._ARRAY).value
            Array = '' if not Array else Array
            if '\n' in Array or ' ' in Array:
               Array = Array.strip().replace(' ', '')
               WS.cell(row=row, column=baseCol+Define._ARRAY, value=Array)

            port.array = Array

            # check InstName
            InstName = WS.cell(row=row, column=baseCol+Define._INSTNAME).value
            InstName = '' if not InstName else InstName
            if '\n' in InstName or ' ' in InstName:
               InstName = InstName.strip().replace(' ', '')
               WS.cell(row=row, column=baseCol+Define._INSTNAME, value=InstName)

            if bool(re.search(r'<\S{1,}>~', InstName)) or bool(re.search(r'~<\S{1,}>', InstName)):
               InstName = InstName.replace('~', Name).replace('<','').replace('>','')

            if InstName.lower() == 'x':
               port.connectType = Define._PRE_DEC
               if   port.direction == Define._PARAM:
                  InstName = m_LstModule.findPortNameByRow(row, i-1)
               elif port.direction == Define._OUTPUT or port.direction == Define._INPUT:
                  InstName = m_LstModule.findInstNameByRow(row, i-1)

            elif InstName == '#':
               port.connectType = Define._SAME_NAME
               InstName = Name
            elif InstName == '':
               port.connectType = Define._FLOAT

            elif bool(re.match(r'\d{1,}\'[a-zA-Z]\d{1,}', InstName)):
               port.connectType = Define._FLOAT

            else : #if InstName != '':
               port.connectType = Define._SPEC_NAME

            port.InstName = InstName
            port.row = row

            if (port.direction != Define._PARAM ) \
               and (port.connectType == Define._SPEC_NAME or \
                     port.connectType == Define._SAME_NAME) \
               and (i != 0):
               if not m_LstModule.containInstNameInWireLst(InstName) and not m_LstModule.containInstNameInTopModulePortLst(InstName):
                  m_LstModule.LstWireDeClaration.append(port)

            m_module.LstPort.append(port)

         # add list port of Top module to LstModule
         if i == 0:
            m_LstModule.LstTopModulePort = m_module.LstPort

         m_LstModule.LstModule.append(m_module)
      return m_LstModule


class Writer():
   def __init__(self, LstModule, FilePath='.'):
      self.LstModule = LstModule
      self.FilePath = FilePath

      name = self.LstModule.LstModule[0].name
      File = open(f'{FilePath}/{name}.sv', 'w')
      File.write(self.WriteModuleDef())

      File.write(self.WriteWireDef())

      for i in range(1, len(self.LstModule.LstModule)):
         File.write(self.WriteModuleInst(i))

      File.write(f'\n`include "{name}_lib.svh"\n')
      File.write(Define.EndDef)
      File.close()



   def WriteWireDef(self):
      String = Define.WireDef
      WireStr = ''
      indentation_0 = ''
      indentation_1 = ' '
      indentation_2 = ' '
      indentation_3 = ' '
      for port in self.LstModule.LstWireDeClaration:
         PortType = 'wire' if port.type == '' else port.type
         WireStr += f'{indentation_0}{PortType}{indentation_1}{port.array}{indentation_2}{port.InstName} ;\n'

      String = String.replace('[WireLst]', WireStr)
      return String


   def WriteModuleDef (self):
      String = Define.ModuleDef
      Module = self.LstModule.LstModule[0]
      String = String.replace('[name]', Module.name)
      ParamStr = ''
      PortStr = ''

      indentation_0 = '   '
      indentation_1 = ' '
      indentation_2 = ' '
      indentation_3 = ' '

      for port in Module.LstPort:
         if port.direction == Define._PARAM:
            if ParamStr != '' :
               ParamStr += ','

            ParamStr +=  f'\n{indentation_0}parameter{indentation_1}{port.array}{indentation_2}{port.name} '

      if ParamStr != '':
         ParamStr += '\n'

      if ParamStr != '':
         String = String.replace('[param]', ParamStr)
      else:
         String = String.replace('#([param])', ParamStr)


      for port in Module.LstPort:
         if   port.direction == Define._INPUT:
            direction = 'input '
         elif port.direction == Define._OUTPUT:
            direction = 'output'

         if PortStr != '':
            PortStr += ',\n'

         PortStr += f'{indentation_0}{direction}{indentation_1}{port.type}{indentation_2}{port.array} {port.name} '

      String = String.replace('[portLst]', PortStr)

      return String


   def WriteModuleInst (self, module_i):
      Module = self.LstModule.LstModule[module_i]
      String = Define.ModuleInst
      String = String.replace('[name]', Module.name)
      String = String.replace('[instName]', Module.name)

      ParamStr = ''
      PortStr = ''

      indentation_0 = '   '
      indentation_1 = ' '
      indentation_2 = ' '
      indentation_3 = ' '

      for port in Module.LstPort:
         if port.direction == Define._PARAM:
            PortType = 'wire' if port.type == '' else port.type
            if ParamStr != '' :
               ParamStr += ','

            ParamStr += f'\n{indentation_0}{port.name}{indentation_2}( {port.InstName} ) '

      if ParamStr != '':
         ParamStr += '\n'

      if ParamStr != '':
         String = String.replace('[param]', ParamStr)
      else:
         String = String.replace('#([param])', ParamStr)


      for port in Module.LstPort:
         if   port.direction == Define._INPUT:
            direction = 'input '
         elif port.direction == Define._OUTPUT:
            direction = 'output'

         if PortStr != '':
            PortStr += f',\n'

         PortStr += f'{indentation_0}.{port.name}{indentation_1}( {port.InstName} ) '
         
      String = String.replace('[portLst]', PortStr)

      return String


def main():
   try:
      xl = XLParser(sys.argv[1])
   except:
      print ('Please input an Excel file')
      os._exit(1)

   try:
      FilePath = sys.argv[2]
   except:
      FilePath = '.'

   for sheet in xl.getLstSheet():
      if sheet.lower() == 'userguide':
         continue
      m_LstModule = xl.parseXL(sheet)
      write = Writer(m_LstModule, FilePath)


if __name__ == '__main__':
   main()